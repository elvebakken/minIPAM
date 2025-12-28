from __future__ import annotations

import os
import base64
import csv
import json
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Dict, Optional, List
from datetime import datetime, timezone

from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Response, Request, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse

from PIL import Image

from .models import (
    LoginRequest, MeResponse, CreateVlanRequest, PatchVlanRequest,
    CreateAssignmentRequest, PatchAssignmentRequest, PatchSettingsRequest,
    ChangePasswordRequest, ChangeUsernameRequest, CreateUserRequest,
    Vlan, Assignment, User
)
from .storage import ensure_files, ensure_admin_user, load_data, save_data, load_users, save_users
from .auth import (
    COOKIE_NAME, cookie_params, create_session_token,
    verify_password, hash_password, require_user, require_role,
    generate_csrf_token, set_csrf_cookie, require_csrf, require_csrf_and_role
)
from .rate_limit import (
    get_client_ip, check_rate_limit, record_failed_attempt, record_successful_login
)
from .ipcalc import parse_network, usable_range, gateway_suggestion, ip_in_subnet, is_network_or_broadcast, next_available_ip
from .audit import append_audit, utcnow_iso, read_audit_logs


def ulid_like(prefix: str) -> str:
    # lightweight unique id: timestamp + random
    import time, secrets
    return f"{prefix}_{int(time.time()*1000)}_{secrets.token_hex(6)}"

DATA_DIR = Path(os.getenv("DATA_DIR", "/data"))
ensure_files(DATA_DIR)

app = FastAPI(title="Mini-IPAM", version="0.1.0")

@app.on_event("startup")
def startup_event():
    """Create admin user on startup if needed."""
    ensure_admin_user(DATA_DIR)


def audit(user: dict, action: str, entity: str, entity_id: str, vlan_id: Optional[str], before: Any, after: Any):
    entry = {
        "ts": utcnow_iso(),
        "user": user.get("u", "unknown"),
        "action": action,
        "entity": entity,
        "entity_id": entity_id,
        "vlan_id": vlan_id,
    }
    if before is not None:
        entry["before"] = before
    if after is not None:
        entry["after"] = after
    append_audit(DATA_DIR / "audit.log", entry)


@app.get("/api/health")
def health():
    return {"ok": True}


# ---------------- AUTH ----------------

@app.post("/api/auth/login")
def login(payload: LoginRequest, request: Request, response: Response):
    # Get client IP for rate limiting
    client_ip = get_client_ip(request)
    username = payload.username.strip() if payload.username and payload.username.strip() else None
    
    # Check rate limits before attempting authentication
    allowed, error_msg = check_rate_limit(DATA_DIR, client_ip, username)
    if not allowed:
        raise HTTPException(status_code=429, detail=error_msg)
    
    users_file = load_users(DATA_DIR)
    user = next((u for u in users_file.users if u.username == payload.username), None)
    
    # Record failed attempt if user doesn't exist or is disabled
    if not user or user.disabled:
        record_failed_attempt(DATA_DIR, client_ip, username)
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Record failed attempt if password is incorrect
    if not verify_password(payload.password, user.password_bcrypt):
        record_failed_attempt(DATA_DIR, client_ip, username)
        raise HTTPException(status_code=401, detail="Invalid credentials")

    # Successful login - clear rate limiting state for this username
    record_successful_login(DATA_DIR, client_ip, username)

    token = create_session_token(user.username, user.role)
    response.set_cookie(COOKIE_NAME, token, **cookie_params())
    
    # Set CSRF token cookie
    csrf_token = generate_csrf_token()
    set_csrf_cookie(response, csrf_token)
    
    return {
        "ok": True,
        "user": {
            "username": user.username,
            "role": user.role,
            "password_change_required": user.password_change_required
        }
    }


@app.post("/api/auth/logout")
def logout(response: Response, _user=Depends(require_csrf)):
    response.delete_cookie(COOKIE_NAME, path="/")
    response.delete_cookie("csrf_token", path="/")
    return {"ok": True}


@app.post("/api/auth/change-password")
def change_password(payload: ChangePasswordRequest, user=Depends(require_csrf)):
    users_file = load_users(DATA_DIR)
    db_user = next((u for u in users_file.users if u.username == user["u"]), None)
    if not db_user or db_user.disabled:
        raise HTTPException(status_code=404, detail="User not found")
    
    if not verify_password(payload.current_password, db_user.password_bcrypt):
        raise HTTPException(status_code=401, detail="Current password is incorrect")
    
    if len(payload.new_password) < 8:
        raise HTTPException(status_code=400, detail="New password must be at least 8 characters")
    
    db_user.password_bcrypt = hash_password(payload.new_password)
    db_user.password_change_required = False
    
    save_users(DATA_DIR, users_file)
    audit(user, "user.password_change", "user", db_user.id, None, None, {"password_changed": True})
    return {"ok": True}


@app.post("/api/auth/change-username")
def change_username(payload: ChangeUsernameRequest, response: Response, user=Depends(require_csrf)):
    users_file = load_users(DATA_DIR)
    db_user = next((u for u in users_file.users if u.username == user["u"]), None)
    if not db_user or db_user.disabled:
        raise HTTPException(status_code=404, detail="User not found")
    
    new_username = payload.new_username.strip()
    if not new_username:
        raise HTTPException(status_code=400, detail="Username cannot be empty")
    
    if len(new_username) < 3:
        raise HTTPException(status_code=400, detail="Username must be at least 3 characters")
    
    # Check if username already exists
    if any(u.username == new_username and u.id != db_user.id for u in users_file.users):
        raise HTTPException(status_code=409, detail="Username already exists")
    
    old_username = db_user.username
    db_user.username = new_username
    
    save_users(DATA_DIR, users_file)
    audit(user, "user.username_change", "user", db_user.id, None, {"old_username": old_username}, {"new_username": new_username})
    
    # Update session token with new username
    token = create_session_token(new_username, db_user.role)
    response.set_cookie(COOKIE_NAME, token, **cookie_params())
    return {"ok": True, "username": new_username}


@app.get("/api/me", response_model=MeResponse)
def me(response: Response, request: Request, user=Depends(require_user)):
    users_file = load_users(DATA_DIR)
    db_user = next((u for u in users_file.users if u.username == user["u"]), None)
    password_change_required = db_user.password_change_required if db_user else False
    
    # Ensure CSRF token cookie is set (refresh if missing)
    if not request.cookies.get("csrf_token"):
        csrf_token = generate_csrf_token()
        set_csrf_cookie(response, csrf_token)
    
    return {
        "username": user["u"],
        "role": user["r"],
        "password_change_required": password_change_required
    }


@app.post("/api/users")
def create_user(payload: CreateUserRequest, user=Depends(require_csrf_and_role({"admin"}))):
    users_file = load_users(DATA_DIR)
    
    # Check if username already exists
    if any(u.username == payload.username for u in users_file.users):
        raise HTTPException(status_code=409, detail="Username already exists")
    
    # Validate username
    username = payload.username.strip()
    if not username:
        raise HTTPException(status_code=400, detail="Username cannot be empty")
    if len(username) < 3:
        raise HTTPException(status_code=400, detail="Username must be at least 3 characters")
    
    # Validate password
    if len(payload.password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters")
    
    # Create new user
    now = utcnow_iso()
    new_user = User(
        id=ulid_like("user"),
        username=username,
        password_bcrypt=hash_password(payload.password),
        role=payload.role,
        created_at=now,
        disabled=False,
        password_change_required=False
    )
    
    users_file.users.append(new_user)
    save_users(DATA_DIR, users_file)
    audit(user, "user.create", "user", new_user.id, None, None, {"username": new_user.username, "role": new_user.role})
    
    return {
        "ok": True,
        "user": {
            "id": new_user.id,
            "username": new_user.username,
            "role": new_user.role
        }
    }


# ---------------- SETTINGS ----------------

@app.get("/api/settings")
def get_settings(user=Depends(require_user)):
    data = load_data(DATA_DIR)
    return data.settings.model_dump()


@app.patch("/api/settings")
def patch_settings(payload: PatchSettingsRequest, user=Depends(require_csrf_and_role({"admin"}))):
    data = load_data(DATA_DIR)
    before = data.settings.model_dump()

    if payload.type_options is not None:
        data.settings.type_options = payload.type_options
    if payload.gateway_default is not None:
        data.settings.gateway_default = payload.gateway_default  # type: ignore
    if payload.reserved_defaults is not None:
        data.settings.reserved_defaults = payload.reserved_defaults

    save_data(DATA_DIR, data)
    audit(user, "settings.update", "settings", "settings", None, before, data.settings.model_dump())
    return {"ok": True, "settings": data.settings.model_dump()}


# ---------------- VLAN HELPERS ----------------

def derive_vlan(data_settings: dict, vlan: Vlan) -> Dict[str, Any]:
    net = parse_network(vlan.subnet_cidr)
    start, end, total = usable_range(net)
    gw = vlan.gateway_ip
    if not gw and data_settings.get("gateway_default") == "first_usable":
        gw = gateway_suggestion(net)
    return {
        "network": str(net.network_address),
        "broadcast": str(net.broadcast_address),
        "usable_start": start,
        "usable_end": end,
        "total_usable": total,
        "gateway_suggested": gateway_suggestion(net),
        "gateway_ip": gw,
    }

def reserved_set(vlan: Vlan, settings: dict) -> set[str]:
    net = parse_network(vlan.subnet_cidr)
    res = set()
    rd = settings.get("reserved_defaults", {})
    if rd.get("reserve_network", True):
        res.add(str(net.network_address))
    if rd.get("reserve_broadcast", True):
        res.add(str(net.broadcast_address))
    # gateway: if set, reserve it; else reserve suggested when enabled
    if rd.get("reserve_gateway", True):
        gw = vlan.gateway_ip or gateway_suggestion(net)
        if gw:
            res.add(gw)
    for r in vlan.reserved_ips:
        res.add(r.ip)
    return res

def used_set(vlan: Vlan) -> set[str]:
    return {a.ip for a in vlan.assignments if not a.archived}


# ---------------- VLAN CRUD ----------------

@app.get("/api/vlans")
def list_vlans(user=Depends(require_user)):
    data = load_data(DATA_DIR)
    out = []
    for v in data.vlans:
        d = derive_vlan(data.settings.model_dump(), v)
        res = reserved_set(v, data.settings.model_dump())
        used = used_set(v)
        out.append({
            "id": v.id,
            "name": v.name,
            "vlan_id": v.vlan_id,
            "subnet_cidr": v.subnet_cidr,
            "gateway_ip": v.gateway_ip or d.get("gateway_ip"),
            "derived": {
                "total_usable": d["total_usable"],
                "reserved": len(res),
                "used": len(used),
                "free": max(d["total_usable"] - len(used) - max(0, (len(res) - 2)), 0)  # approximate; UI uses counts mainly
            }
        })
    return out


@app.post("/api/vlans")
def create_vlan(payload: CreateVlanRequest, user=Depends(require_csrf_and_role({"admin", "readwrite"}))):
    data = load_data(DATA_DIR)
    # validate cidr
    try:
        net = parse_network(payload.subnet_cidr)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid CIDR: {e}")

    # Forbid /31 and /32 subnets as they have no usable hosts
    if net.prefixlen >= 31:
        raise HTTPException(status_code=400, detail=f"Subnet /{net.prefixlen} has no usable hosts. /31 and /32 subnets are not supported for VLAN creation.")

    now = utcnow_iso()
    v = Vlan(
        id=ulid_like("vlan"),
        name=payload.name,
        vlan_id=payload.vlan_id,
        subnet_cidr=str(net),
        gateway_ip=gateway_suggestion(net) if data.settings.gateway_default == "first_usable" else None,
        reserved_ips=[],
        assignments=[],
        created_at=now,
        updated_at=now,
    )
    data.vlans.append(v)
    save_data(DATA_DIR, data)
    audit(user, "vlan.create", "vlan", v.id, v.id, None, {"name": v.name, "subnet_cidr": v.subnet_cidr})
    return v.model_dump()


@app.get("/api/vlans/{vlan_id}")
def get_vlan(vlan_id: str, user=Depends(require_user)):
    data = load_data(DATA_DIR)
    v = next((x for x in data.vlans if x.id == vlan_id), None)
    if not v:
        raise HTTPException(status_code=404, detail="VLAN not found")
    d = derive_vlan(data.settings.model_dump(), v)
    res = sorted(list(reserved_set(v, data.settings.model_dump())))
    used = sorted(list(used_set(v)))
    return {
        **v.model_dump(),
        "derived": d,
        "reserved_effective": res,
        "used_effective": used
    }


@app.patch("/api/vlans/{vlan_id}")
def patch_vlan(vlan_id: str, payload: PatchVlanRequest, user=Depends(require_csrf_and_role({"admin", "readwrite"}))):
    data = load_data(DATA_DIR)
    v = next((x for x in data.vlans if x.id == vlan_id), None)
    if not v:
        raise HTTPException(status_code=404, detail="VLAN not found")
    before = v.model_dump()

    if payload.name is not None:
        v.name = payload.name
    if payload.vlan_id is not None:
        v.vlan_id = payload.vlan_id
    if payload.subnet_cidr is not None:
        try:
            net = parse_network(payload.subnet_cidr)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid CIDR: {e}")
        # Forbid /31 and /32 subnets as they have no usable hosts
        if net.prefixlen >= 31:
            raise HTTPException(status_code=400, detail=f"Subnet /{net.prefixlen} has no usable hosts. /31 and /32 subnets are not supported for VLAN creation.")
        v.subnet_cidr = str(net)
        # optional: reset gateway suggestion if not set
        if not v.gateway_ip and data.settings.gateway_default == "first_usable":
            v.gateway_ip = gateway_suggestion(net)
    if payload.gateway_ip is not None:
        if payload.gateway_ip != "" and not ip_in_subnet(payload.gateway_ip, v.subnet_cidr):
            raise HTTPException(status_code=400, detail="Gateway IP must be inside subnet")
        v.gateway_ip = payload.gateway_ip or None

    v.updated_at = utcnow_iso()
    save_data(DATA_DIR, data)
    audit(user, "vlan.update", "vlan", v.id, v.id, {"name": before["name"], "subnet_cidr": before["subnet_cidr"]},
          {"name": v.name, "subnet_cidr": v.subnet_cidr})
    return {"ok": True}


@app.delete("/api/vlans/{vlan_id}")
def delete_vlan(vlan_id: str, user=Depends(require_csrf_and_role({"admin", "readwrite"}))):
    data = load_data(DATA_DIR)
    idx = next((i for i, x in enumerate(data.vlans) if x.id == vlan_id), None)
    if idx is None:
        raise HTTPException(status_code=404, detail="VLAN not found")
    before = data.vlans[idx].model_dump()
    data.vlans.pop(idx)
    save_data(DATA_DIR, data)
    audit(user, "vlan.delete", "vlan", vlan_id, vlan_id, {"name": before["name"]}, None)
    return {"ok": True}


# ---------------- ASSIGNMENTS ----------------

@app.get("/api/vlans/{vlan_id}/next-available")
def api_next_available(vlan_id: str, user=Depends(require_user)):
    data = load_data(DATA_DIR)
    v = next((x for x in data.vlans if x.id == vlan_id), None)
    if not v:
        raise HTTPException(status_code=404, detail="VLAN not found")

    res = reserved_set(v, data.settings.model_dump())
    used = used_set(v)
    ip = next_available_ip(v.subnet_cidr, used, res)
    return {"ip": ip}


def normalize_and_validate_assignment(v: Vlan, data_settings: dict, ip: str, assignment_id: Optional[str] = None):
    # cidr-aware check
    if not ip_in_subnet(ip, v.subnet_cidr):
        raise HTTPException(status_code=400, detail="IP is outside VLAN subnet")

    # Block network and broadcast addresses via CIDR logic (not string matching)
    nb_check = is_network_or_broadcast(ip, v.subnet_cidr)
    if nb_check["is_network"]:
        raise HTTPException(status_code=400, detail="IP is the network address for this subnet")
    if nb_check["is_broadcast"]:
        raise HTTPException(status_code=400, detail="IP is the broadcast address for this subnet")

    # Check if subnet has usable hosts (for /31 and /32)
    net = parse_network(v.subnet_cidr)
    if net.prefixlen >= 31:
        raise HTTPException(status_code=400, detail="Subnet has no usable hosts (/31 and /32 subnets cannot have assignments)")

    # prevent duplicates (excluding self on patch)
    for a in v.assignments:
        if a.archived:
            continue
        if a.ip == ip and (assignment_id is None or a.id != assignment_id):
            raise HTTPException(status_code=409, detail="Duplicate IP in this VLAN")

    # reserved block
    res = reserved_set(v, data_settings)
    if ip in res:
        # Allow reserved only if it's explicitly listed as a custom reserved with reason AND not used by assignment
        raise HTTPException(status_code=400, detail="IP is reserved in this VLAN")


@app.post("/api/vlans/{vlan_id}/assignments")
def create_assignment(vlan_id: str, payload: CreateAssignmentRequest, user=Depends(require_csrf_and_role({"admin", "readwrite"}))):
    data = load_data(DATA_DIR)
    v = next((x for x in data.vlans if x.id == vlan_id), None)
    if not v:
        raise HTTPException(status_code=404, detail="VLAN not found")

    normalize_and_validate_assignment(v, data.settings.model_dump(), payload.ip)

    now = utcnow_iso()
    a = Assignment(
        id=ulid_like("asgn"),
        ip=payload.ip,
        hostname=payload.hostname,
        type=payload.type,
        tags=payload.tags,
        notes=payload.notes,
        icon=payload.icon,
        archived=False,
        created_at=now,
        updated_at=now,
    )
    v.assignments.append(a)
    v.updated_at = now

    save_data(DATA_DIR, data)
    audit(user, "assignment.create", "assignment", a.id, v.id, None, {"ip": a.ip, "hostname": a.hostname, "type": a.type, "tags": a.tags})
    return a.model_dump()


@app.patch("/api/vlans/{vlan_id}/assignments/{assignment_id}")
def patch_assignment(vlan_id: str, assignment_id: str, payload: PatchAssignmentRequest, user=Depends(require_csrf_and_role({"admin", "readwrite"}))):
    data = load_data(DATA_DIR)
    v = next((x for x in data.vlans if x.id == vlan_id), None)
    if not v:
        raise HTTPException(status_code=404, detail="VLAN not found")
    a = next((x for x in v.assignments if x.id == assignment_id), None)
    if not a:
        raise HTTPException(status_code=404, detail="Assignment not found")
    before = {"ip": a.ip, "hostname": a.hostname, "type": a.type, "tags": a.tags, "archived": a.archived}

    if payload.ip is not None:
        normalize_and_validate_assignment(v, data.settings.model_dump(), payload.ip, assignment_id=a.id)
        a.ip = payload.ip
    if payload.hostname is not None:
        a.hostname = payload.hostname
    if payload.type is not None:
        a.type = payload.type
    if payload.tags is not None:
        a.tags = payload.tags
    if payload.notes is not None:
        a.notes = payload.notes
    if payload.icon is not None:
        a.icon = payload.icon
    if payload.archived is not None:
        a.archived = payload.archived

    a.updated_at = utcnow_iso()
    v.updated_at = a.updated_at

    save_data(DATA_DIR, data)
    audit(user, "assignment.update", "assignment", a.id, v.id, before, {"ip": a.ip, "hostname": a.hostname, "type": a.type, "tags": a.tags, "archived": a.archived})
    return {"ok": True}


@app.delete("/api/vlans/{vlan_id}/assignments/{assignment_id}")
def delete_assignment(vlan_id: str, assignment_id: str, user=Depends(require_csrf_and_role({"admin", "readwrite"}))):
    data = load_data(DATA_DIR)
    v = next((x for x in data.vlans if x.id == vlan_id), None)
    if not v:
        raise HTTPException(status_code=404, detail="VLAN not found")
    idx = next((i for i, x in enumerate(v.assignments) if x.id == assignment_id), None)
    if idx is None:
        raise HTTPException(status_code=404, detail="Assignment not found")
    before = v.assignments[idx].model_dump()
    v.assignments.pop(idx)
    v.updated_at = utcnow_iso()
    save_data(DATA_DIR, data)
    audit(user, "assignment.delete", "assignment", assignment_id, v.id, {"ip": before.get("ip")}, None)
    return {"ok": True}


# ---------------- ICON UPLOAD ----------------

ICONS_DIR = Path(__file__).parent.parent / "icons"

@app.get("/api/icons/list")
def list_icons(user=Depends(require_user)):
    """List available predefined icons."""
    if not ICONS_DIR.exists():
        return {"icons": []}
    
    icons = []
    for file in sorted(ICONS_DIR.glob("*.png")):
        icons.append({
            "name": file.stem,
            "filename": file.name
        })
    return {"icons": icons}

@app.get("/api/icons/{icon_name}")
def get_icon(icon_name: str, user=Depends(require_role({"admin", "readwrite"}))):
    """Load and normalize a predefined icon."""
    # Security: only allow PNG files, prevent path traversal
    if not icon_name.endswith(".png") or "/" in icon_name or "\\" in icon_name:
        raise HTTPException(status_code=400, detail="Invalid icon name")
    
    icon_path = ICONS_DIR / icon_name
    if not icon_path.exists() or not icon_path.is_file():
        raise HTTPException(status_code=404, detail="Icon not found")
    
    try:
        with open(icon_path, "rb") as f:
            raw = f.read()
        
        img = Image.open(BytesIO(raw)).convert("RGBA")
        # center-crop to square
        w, h = img.size
        side = min(w, h)
        left = (w - side) // 2
        top = (h - side) // 2
        img = img.crop((left, top, left + side, top + side))
        img = img.resize((256, 256), Image.LANCZOS)

        out = BytesIO()
        img.save(out, format="PNG", optimize=True)
        png_bytes = out.getvalue()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid image: {e}")

    b64 = base64.b64encode(png_bytes).decode("ascii")
    return {"mime_type": "image/png", "data_base64": b64}

@app.post("/api/icons/normalize")
def normalize_icon(user=Depends(require_csrf_and_role({"admin", "readwrite"})), file: UploadFile = File(...)):
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Only images are allowed")

    raw = file.file.read()
    if len(raw) > 2_000_000:
        raise HTTPException(status_code=400, detail="Image too large (max 2MB)")

    try:
        img = Image.open(BytesIO(raw)).convert("RGBA")
        # center-crop to square
        w, h = img.size
        side = min(w, h)
        left = (w - side) // 2
        top = (h - side) // 2
        img = img.crop((left, top, left + side, top + side))
        img = img.resize((256, 256), Image.LANCZOS)

        out = BytesIO()
        img.save(out, format="PNG", optimize=True)
        png_bytes = out.getvalue()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid image: {e}")

    b64 = base64.b64encode(png_bytes).decode("ascii")
    return {"mime_type": "image/png", "data_base64": b64}

@app.post("/api/icons/upload-multiple")
async def upload_multiple_icons(user=Depends(require_csrf_and_role({"admin"})), files: list[UploadFile] = File(...)):
    """Upload multiple icons to the icons directory (admin only)."""
    if not ICONS_DIR.exists():
        ICONS_DIR.mkdir(parents=True, exist_ok=True)
    
    uploaded = []
    errors = []
    
    for file in files:
        if not file.content_type or not file.content_type.startswith("image/"):
            errors.append({"filename": file.filename, "error": "Only images are allowed"})
            continue
        
        try:
            raw = await file.read()
            if len(raw) > 2_000_000:
                errors.append({"filename": file.filename, "error": "Image too large (max 2MB)"})
                continue
            
            # Validate and normalize image
            img = Image.open(BytesIO(raw)).convert("RGBA")
            w, h = img.size
            side = min(w, h)
            left = (w - side) // 2
            top = (h - side) // 2
            img = img.crop((left, top, left + side, top + side))
            img = img.resize((256, 256), Image.LANCZOS)
            
            # Generate safe filename
            safe_name = "".join(c for c in file.filename if c.isalnum() or c in "._- ").strip()
            if not safe_name:
                safe_name = "icon"
            if not safe_name.endswith(".png"):
                safe_name = safe_name.rsplit(".", 1)[0] + ".png"
            
            # Ensure unique filename
            base_name = safe_name.rsplit(".", 1)[0]
            counter = 1
            final_path = ICONS_DIR / safe_name
            while final_path.exists():
                final_path = ICONS_DIR / f"{base_name}_{counter}.png"
                counter += 1
            
            # Save as PNG
            img.save(final_path, format="PNG", optimize=True)
            uploaded.append({"filename": final_path.name, "name": base_name})
            audit(user, "icon.upload", "icon", final_path.name, None, None, {"filename": final_path.name})
        except Exception as e:
            errors.append({"filename": file.filename, "error": str(e)})
    
    return {
        "uploaded": uploaded,
        "errors": errors,
        "total": len(files),
        "success_count": len(uploaded)
    }

@app.delete("/api/icons/{icon_name}")
def delete_icon(icon_name: str, user=Depends(require_csrf_and_role({"admin"}))):
    """Delete an icon from the icons directory (admin only)."""
    # Security: only allow PNG files, prevent path traversal
    if not icon_name.endswith(".png") or "/" in icon_name or "\\" in icon_name:
        raise HTTPException(status_code=400, detail="Invalid icon name")
    
    icon_path = ICONS_DIR / icon_name
    if not icon_path.exists() or not icon_path.is_file():
        raise HTTPException(status_code=404, detail="Icon not found")
    
    try:
        icon_path.unlink()
        audit(user, "icon.delete", "icon", icon_name, None, {"filename": icon_name}, None)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete icon: {e}")


# ---------------- EXPORT ----------------

@app.get("/api/export/data")
def export_data(user=Depends(require_role({"admin", "readwrite", "readonly"}))):
    path = DATA_DIR / "data.json"
    if not path.exists():
        raise HTTPException(status_code=404, detail="data.json missing")
    return FileResponse(str(path), media_type="application/json", filename="data.json")


@app.get("/api/vlans/{vlan_id}/assignments/export")
def export_assignments(
    vlan_id: str,
    format: str = Query("csv", regex="^(csv|json|excel)$"),
    search: Optional[str] = None,
    type_filter: Optional[str] = None,
    user=Depends(require_user)
):
    """Export assignments from a VLAN in CSV, JSON, or Excel format with optional filtering."""
    data = load_data(DATA_DIR)
    v = next((x for x in data.vlans if x.id == vlan_id), None)
    if not v:
        raise HTTPException(status_code=404, detail="VLAN not found")
    
    # Filter assignments
    assignments = [a for a in v.assignments if not a.archived]
    
    # Apply search filter
    if search:
        search_lower = search.lower()
        assignments = [
            a for a in assignments
            if search_lower in a.ip.lower() or
               search_lower in (a.hostname or "").lower() or
               search_lower in a.type.lower() or
               search_lower in (a.notes or "").lower() or
               any(search_lower in tag.lower() for tag in (a.tags or []))
        ]
    
    # Apply type filter
    if type_filter and type_filter != "all":
        assignments = [a for a in assignments if a.type == type_filter]
    
    # Sort by IP
    assignments.sort(key=lambda x: x.ip)
    
    # Prepare data with VLAN context
    export_data = []
    for a in assignments:
        export_data.append({
            "vlan_name": v.name,
            "vlan_id": v.vlan_id,
            "subnet_cidr": v.subnet_cidr,
            "ip": a.ip,
            "hostname": a.hostname,
            "type": a.type,
            "tags": ", ".join(a.tags) if a.tags else "",
            "notes": a.notes,
            "created_at": a.created_at,
            "updated_at": a.updated_at,
        })
    
    if format == "json":
        json_str = json.dumps(export_data, indent=2)
        return Response(
            content=json_str,
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="assignments_{v.name}_{vlan_id}.json"'}
        )
    
    elif format == "csv":
        output = StringIO()
        if export_data:
            writer = csv.DictWriter(output, fieldnames=export_data[0].keys())
            writer.writeheader()
            writer.writerows(export_data)
        
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="assignments_{v.name}_{vlan_id}.csv"'}
        )
    
    elif format == "excel":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Assignments"
            
            # Headers
            if export_data:
                headers = list(export_data[0].keys())
                ws.append(headers)
                
                # Style header row
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                
                # Add data rows
                for row_data in export_data:
                    ws.append([row_data.get(h, "") for h in headers])
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="assignments_{v.name}_{vlan_id}.xlsx"'}
            )
        except ImportError:
            raise HTTPException(status_code=500, detail="Excel export requires openpyxl library")


@app.post("/api/vlans/{vlan_id}/assignments/import")
def import_assignments(
    vlan_id: str,
    file: UploadFile = File(...),
    user=Depends(require_csrf_and_role({"admin", "readwrite"}))
):
    """Import assignments from CSV, JSON, or Excel file."""
    data = load_data(DATA_DIR)
    v = next((x for x in data.vlans if x.id == vlan_id), None)
    if not v:
        raise HTTPException(status_code=404, detail="VLAN not found")
    
    file_content = file.file.read()
    filename = file.filename or ""
    
    imported = []
    errors = []
    
    try:
        # Determine file type
        if filename.endswith('.json') or file.content_type == 'application/json':
            # JSON import
            try:
                json_data = json.loads(file_content.decode('utf-8'))
                if isinstance(json_data, list):
                    rows = json_data
                else:
                    rows = [json_data]
            except json.JSONDecodeError as e:
                raise HTTPException(status_code=400, detail=f"Invalid JSON: {e}")
        
        elif filename.endswith('.csv') or file.content_type == 'text/csv':
            # CSV import
            csv_content = file_content.decode('utf-8')
            reader = csv.DictReader(StringIO(csv_content))
            rows = list(reader)
        
        elif filename.endswith(('.xlsx', '.xls')) or 'spreadsheet' in (file.content_type or ''):
            # Excel import
            try:
                from openpyxl import load_workbook
                wb = load_workbook(BytesIO(file_content))
                ws = wb.active
                
                # Read headers
                headers = [cell.value for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=False):
                    row_dict = {}
                    for i, cell in enumerate(row):
                        if i < len(headers) and headers[i]:
                            row_dict[headers[i]] = cell.value
                    if any(row_dict.values()):  # Skip empty rows
                        rows.append(row_dict)
            except ImportError:
                raise HTTPException(status_code=500, detail="Excel import requires openpyxl library")
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e}")
        
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Use CSV, JSON, or Excel.")
        
        # Process rows
        now = utcnow_iso()
        for idx, row in enumerate(rows, start=2):  # Start at 2 for Excel compatibility
            try:
                # Extract IP (required)
                ip = str(row.get('ip', row.get('IP', ''))).strip()
                if not ip:
                    errors.append({"row": idx, "error": "IP address is required"})
                    continue
                
                # Validate IP is in subnet
                if not ip_in_subnet(ip, v.subnet_cidr):
                    errors.append({"row": idx, "ip": ip, "error": f"IP {ip} is outside VLAN subnet {v.subnet_cidr}"})
                    continue
                
                # Block network and broadcast addresses via CIDR logic
                nb_check = is_network_or_broadcast(ip, v.subnet_cidr)
                if nb_check["is_network"]:
                    errors.append({"row": idx, "ip": ip, "error": f"IP {ip} is the network address for this subnet"})
                    continue
                if nb_check["is_broadcast"]:
                    errors.append({"row": idx, "ip": ip, "error": f"IP {ip} is the broadcast address for this subnet"})
                    continue
                
                # Check if subnet has usable hosts (for /31 and /32)
                net = parse_network(v.subnet_cidr)
                if net.prefixlen >= 31:
                    errors.append({"row": idx, "ip": ip, "error": f"Subnet has no usable hosts (/31 and /32 subnets cannot have assignments)"})
                    continue
                
                # Check for duplicates
                if any(a.ip == ip and not a.archived for a in v.assignments):
                    errors.append({"row": idx, "ip": ip, "error": f"IP {ip} already exists in this VLAN"})
                    continue
                
                # Check if reserved
                res = reserved_set(v, data.settings.model_dump())
                if ip in res:
                    errors.append({"row": idx, "ip": ip, "error": f"IP {ip} is reserved in this VLAN"})
                    continue
                
                # Extract other fields
                hostname = str(row.get('hostname', row.get('Hostname', ''))).strip()
                assignment_type = str(row.get('type', row.get('Type', 'server'))).strip() or 'server'
                
                # Handle tags (can be comma-separated string or list)
                tags_raw = row.get('tags', row.get('Tags', ''))
                if isinstance(tags_raw, list):
                    tags = [str(t).strip() for t in tags_raw if t]
                else:
                    tags = [t.strip() for t in str(tags_raw).split(',') if t.strip()]
                
                notes = str(row.get('notes', row.get('Notes', ''))).strip()
                
                # Create assignment
                a = Assignment(
                    id=ulid_like("asgn"),
                    ip=ip,
                    hostname=hostname,
                    type=assignment_type,
                    tags=tags,
                    notes=notes,
                    icon=None,
                    archived=False,
                    created_at=now,
                    updated_at=now,
                )
                
                v.assignments.append(a)
                imported.append({"ip": ip, "hostname": hostname})
                
            except Exception as e:
                errors.append({"row": idx, "error": f"Error processing row: {str(e)}"})
        
        if imported:
            v.updated_at = now
            save_data(DATA_DIR, data)
            audit(user, "assignment.import", "assignment", f"bulk_{vlan_id}", v.id, None, {"count": len(imported)})
        
        return {
            "ok": True,
            "imported": len(imported),
            "errors": len(errors),
            "imported_items": imported,
            "error_details": errors
        }
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Import failed: {str(e)}")


# ---------------- AUDIT LOGS ----------------

@app.get("/api/audit-logs")
def get_audit_logs(
    user=Depends(require_user),
    user_filter: Optional[str] = None,
    action_filter: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 1000
):
    """Get audit logs with optional filtering."""
    audit_path = DATA_DIR / "audit.log"
    entries = read_audit_logs(audit_path, user_filter, action_filter, date_from, date_to, limit)
    return {"entries": entries}


# Serve static UI - mount at the end so API routes take precedence
static_dir = Path(__file__).parent / "static"
app.mount("/", StaticFiles(directory=static_dir, html=True), name="static")

# Serve icons folder
if ICONS_DIR.exists():
    app.mount("/icons", StaticFiles(directory=ICONS_DIR), name="icons")
